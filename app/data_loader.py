"""Load shipment records from Excel or published JSON."""
from __future__ import annotations

import json
from datetime import date, datetime
from pathlib import Path
from typing import Any, Dict, List

from openpyxl import load_workbook

from app.config import (
    DATE_FIELDS,
    JSON_FILE,
    LEGACY_WORKBOOK,
    OUTPUT_SHEET_NAME,
    PUBLIC_FIELDS,
    WORKBOOK_FILE,
)


def _text(value: Any) -> str:
    if value is None or str(value).strip().upper() in {"", "NA", "N/A", "NAN", "NONE"}:
        return ""
    return str(value).strip()


def _date(value: Any) -> str:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = _text(value)
    if not text:
        return ""
    for fmt in ("%m/%d/%Y", "%Y-%m-%d", "%d/%m/%Y"):
        try:
            return datetime.strptime(text, fmt).date().isoformat()
        except ValueError:
            pass
    return ""


def _forwarder(source: str) -> str:
    source = _text(source).lower()
    for label in ("ScanGlobal", "CargoMar", "Sinpex", "Liyana", "GoComet"):
        if label.lower() in source:
            return label
    return "Detailed Tracking" if "detailed" in source else "Other"


def _arrival_status(record: Dict[str, str]) -> str:
    if record.get("Vessel Arrived") or record.get("Container Delivered"):
        return "Arrived / Delivered"
    return "In Transit"


def enrich(record: Dict[str, str]) -> Dict[str, str]:
    record["Forwarder"] = _forwarder(record.get("Source File", ""))
    record["Arrived/Delivered"] = _arrival_status(record)
    return record


def slim(record: Dict[str, str]) -> Dict[str, str]:
    out = {field: record.get(field, "") for field in PUBLIC_FIELDS}
    out["Forwarder"] = record.get("Forwarder", "")
    out["Arrived/Delivered"] = record.get("Arrived/Delivered", "")
    return out


def load_workbook_records(workbook_path: Path) -> List[Dict[str, str]]:
    if not workbook_path.exists():
        raise FileNotFoundError(f"Workbook not found: {workbook_path}")

    wb = load_workbook(workbook_path, read_only=True, data_only=True)
    if OUTPUT_SHEET_NAME not in wb.sheetnames:
        raise ValueError(f"Sheet '{OUTPUT_SHEET_NAME}' not found in {workbook_path.name}")

    ws = wb[OUTPUT_SHEET_NAME]
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    records: List[Dict[str, str]] = []

    for values in ws.iter_rows(min_row=2, values_only=True):
        record = {
            str(field): (_date(value) if field in DATE_FIELDS else _text(value))
            for field, value in zip(headers, values)
            if field
        }
        if record.get("Container Number"):
            records.append(enrich(record))

    wb.close()
    return records


def resolve_workbook() -> Path:
    if WORKBOOK_FILE.exists():
        return WORKBOOK_FILE
    if LEGACY_WORKBOOK.exists():
        return LEGACY_WORKBOOK
    raise FileNotFoundError(
        "No consolidated workbook found. Upload source files via /admin or run ETL first."
    )


def load_records(*, public_only: bool = True) -> List[Dict[str, str]]:
    if JSON_FILE.exists():
        payload = json.loads(JSON_FILE.read_text(encoding="utf-8"))
        records = payload.get("records", [])
        return records

    records = load_workbook_records(resolve_workbook())
    return [slim(r) for r in records] if public_only else records


def build_payload(records: List[Dict[str, str]], workbook: Path, version: str) -> Dict[str, Any]:
    modified = datetime.fromtimestamp(workbook.stat().st_mtime).strftime("%b %d, %Y %I:%M %p")
    return {
        "records": records,
        "workbook": workbook.name,
        "modified": modified,
        "version": version,
        "rowCount": len(records),
    }
