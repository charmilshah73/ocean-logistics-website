"""Run v10_pkg main.py ETL without modifying v10_pkg."""
from __future__ import annotations

import os
import sys
from pathlib import Path
from typing import Any, Dict, List

from app.config import INPUT, STAGING, V10_PKG


def run_etl(input_dir: Path | None = None, output_dir: Path | None = None) -> int:
    """Execute consolidation; returns process exit code."""
    input_path = input_dir or INPUT
    output_path = output_dir or (STAGING / "output")
    output_path.mkdir(parents=True, exist_ok=True)

    os.environ["OCEAN_DSR_INPUT_DIR"] = str(input_path)
    os.environ["OCEAN_DSR_OUTPUT_DIR"] = str(output_path)

    if str(V10_PKG) not in sys.path:
        sys.path.insert(0, str(V10_PKG))

    # Reload config + main so env vars take effect.
    for mod_name in ("config", "main"):
        if mod_name in sys.modules:
            del sys.modules[mod_name]

    import main  # noqa: WPS433 — loaded from v10_pkg via sys.path

    return int(main.main())


def collect_warnings(output_dir: Path) -> List[str]:
    """Read validation warnings from audit sheet if present."""
    warnings: List[str] = []
    workbook = output_dir / "Consolidated_Ocean_DSR.xlsx"
    if not workbook.exists():
        return warnings

    try:
        from openpyxl import load_workbook

        wb = load_workbook(workbook, read_only=True, data_only=True)
        if "Validation_Report" not in wb.sheetnames:
            wb.close()
            return warnings
        ws = wb["Validation_Report"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row and row[0] == "Warning" and row[4]:
                warnings.append(str(row[4]))
        wb.close()
    except Exception as exc:
        warnings.append(f"Could not read validation report: {exc}")
    return warnings


def etl_summary(output_dir: Path) -> Dict[str, Any]:
    workbook = output_dir / "Consolidated_Ocean_DSR.xlsx"
    row_count = 0
    if workbook.exists():
        from openpyxl import load_workbook

        from app.config import OUTPUT_SHEET_NAME

        wb = load_workbook(workbook, read_only=True, data_only=True)
        if OUTPUT_SHEET_NAME in wb.sheetnames:
            row_count = max(0, (wb[OUTPUT_SHEET_NAME].max_row or 1) - 1)
        wb.close()
    return {
        "workbook": str(workbook),
        "rowCount": row_count,
        "warnings": collect_warnings(output_dir),
    }
