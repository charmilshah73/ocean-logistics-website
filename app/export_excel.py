"""Build Excel downloads for filtered dashboard views."""
from __future__ import annotations

import io
from datetime import datetime
from typing import Any, Dict, List

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


def build_detail_workbook(records: List[Dict[str, Any]], columns: List[str]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Detailed Report"

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(bold=True, color="FFFFFF")

    for col_idx, name in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.fill = header_fill
        cell.font = header_font

    for row_idx, record in enumerate(records, start=2):
        for col_idx, name in enumerate(columns, start=1):
            ws.cell(row=row_idx, column=col_idx, value=record.get(name, ""))

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}{max(1, len(records) + 1)}"

    for col_idx, name in enumerate(columns, start=1):
        width = min(42, max(len(name), 12))
        for row_idx in range(2, min(len(records) + 2, 200)):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val is not None:
                width = max(width, min(len(str(val)), 42))
        ws.column_dimensions[get_column_letter(col_idx)].width = width + 2

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def detail_filename() -> str:
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return f"Ocean_Detailed_Report_{stamp}.xlsx"
