from __future__ import annotations

import re
import os
import sys
import tempfile
import warnings
import zipfile
import xml.etree.ElementTree as ET
from collections import defaultdict, OrderedDict
from copy import copy
from datetime import datetime, date, timedelta
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple, Set

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
import xlsxwriter

from config import INPUT_DIR, OUTPUT_DIR, OUTPUT_FILE, OUTPUT_SHEET_NAME, OPEN_ORDER_SHEET_NAME, DATE_FORMAT, CREATE_AUDIT_SHEETS

warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl.styles.stylesheet")

NA = "NA"
NA_VALUES = {"", "NA", "N/A", "NONE", "NULL", "-", "--", "NIL", "NAN"}
MAX_TRAILING_BLANK_ROWS = 100

_OPEN_INPUT_WORKBOOKS: List[Any] = []


def open_input_workbook(path: Path):
    """Open a read-only workbook and register it for exception-safe cleanup."""
    workbook = load_workbook(path, read_only=True, data_only=True)
    _OPEN_INPUT_WORKBOOKS.append(workbook)
    return workbook


def close_input_workbook(workbook) -> None:
    """Close a workbook immediately; repeated cleanup remains harmless."""
    try:
        workbook.close()
    finally:
        if workbook in _OPEN_INPUT_WORKBOOKS:
            _OPEN_INPUT_WORKBOOKS.remove(workbook)


def close_all_input_workbooks() -> None:
    """Release every workbook still open after an exception."""
    while _OPEN_INPUT_WORKBOOKS:
        workbook = _OPEN_INPUT_WORKBOOKS.pop()
        try:
            workbook.close()
        except Exception:
            pass

DATE_KEYWORDS = [
    "DATE", "ETA", "ETD", "ATD", "ATA", "ARRIVED", "DEPARTED", "LOADED", "INGATED", "OUTGATED",
    "RELEASED", "DELIVERED", "PICK UP", "PICKUP", "CUT OFF", "CUTOFF", "GATE", "DISCHARGED",
    "FILED", "LFD", "DEPARTURE", "BOOKING DT", "INVOICE DT", "DLV DATE", "CARGO READY", "CREATED AT", "LAST UPDATE"
]

EXTRA_HEADERS_KEEP = [
    "SR NO", "Invoice Dt", "Pick Up Location", "Scheduled Pick up Date", "Detention Days", "Vehicle type", "Vehicle No",
    "Actual SI Cut off Dt", "Si Filed Dt", "Gate / CY Cut Off", "Actual Gate In Dt &Time", "Shipping Bill No", "Seal No",
    "Second Vessel", "Voy_2", "Final Vessel", "Voy_3", "2nd Etd", "3rd Etd", "T/S or Direct", "T/S Vessel", "Etd from T/S",
    "ACTUAL TRANSIT TIME", "Current Status",
]
OPEN_ORDER_OUTPUT_COLUMNS = ["Dlv Date", "Buyer", "Vendor", "Class"]

# Source header normalized -> target header from Liyana template / extra columns
ALIASES = {
    "SGL FILE #": "Lane number",
    "LIYANA LANE NUMBER": "Lane number",
    "LANE NUMBER": "Lane number",
    "CUSTOMER NUMBER": "Customer Number",
    "SHIPPER": "Shipper Name",
    "SHIPPER NAME": "Shipper Name",
    "CONSIGNEE": "Consignee Name",
    "CONSIGNEE NAME": "Consignee Name",
    "DESTINATION PLANT": "Delivery Location",
    "DELIVERY LOCATION": "Delivery Location",
    "PLANT OTHER DATA": "Delivery Location",
    "BOOKING NO": "BOOKING NO",
    "BOOKING NUMBER": "BOOKING NO",
    "CONTAINER NO": "Container Number",
    "CONTAINER NUMBER": "Container Number",
    "CONTAINER": "Container Number",
    "CONTAINER NUMBER IF OCEAN SHIPMENT": "Container Number",
    "TRACKING NUMBER": "Container Number",
    "BL NO": "Steamship Line Bill Of Lading #",
    "B L NO": "Steamship Line Bill Of Lading #",
    "BILL OF LADING": "Steamship Line Bill Of Lading #",
    "STEAMSHIP LINE BILL OF LADING #": "Steamship Line Bill Of Lading #",
    "BOOKING REQUEST DATE": "Booking Request Date",
    "BOOKING DT": "Booking Confirmation Date",
    "BOOKING CONFIRMATION DATE": "Booking Confirmation Date",
    "CARGO READY DATE": "Cargo readiness date",
    "CARGO READINESS DATE": "Cargo readiness date",
    "ACTUAL PICK UP DATE VEHICLE PLACE DATE TIME": "Vehicle Placement date",
    "VEHICLE PLACEMENT DATE": "Vehicle Placement date",
    "VEHICLE RELEASED DATE TIME": "Cargo actual pick up date",
    "CARGO ACTUAL PICK UP DATE": "Cargo actual pick up date",
    "FIRST VESSEL": "Vessel Name",
    "VESSEL NAME": "Vessel Name",
    "ORIGIN DEPARTURE VESSEL NAME": "Vessel Name",
    "VOY": "Voyage No-",
    "VOYAGE NO": "Voyage No-",
    "VOYAGE NO #": "Voyage No-",
    "VOYAGE NO-": "Voyage No-",
    "HBL NUMBER": "HBL Number",
    "HBL NO": "HBL Number",
    "FCL LCL": "MOT",
    "FCL / LCL": "MOT",
    "FCL/LCL": "MOT",
    "MOT": "MOT",
    "INCOTERM": "Incoterm",
    "INCOTERMS": "Incoterm",
    "PO NUMBER": "PO#",
    "PO #": "PO#",
    "PO#": "PO#",
    "PO NO": "PO#",
    "PO NBR": "PO#",
    "REFERENCE": "PO#",
    "CONTAINER TYPE": "CTR SIZE / LCL",
    "CTR SIZE LCL": "CTR SIZE / LCL",
    "CTR SIZE / LCL": "CTR SIZE / LCL",
    "WEIGHT": "Actual Weight",
    "ACTUAL WEIGHT": "Actual Weight",
    "CBM": "Volume(CBM)",
    "VOLUME CBM": "Volume(CBM)",
    "VOLUME(CBM)": "Volume(CBM)",
    "TRANS SHIPMENT PORT": "Transshipment port",
    "TRANSSHIPMENT PORT": "Transshipment port",
    "TRANSIT PORT": "Transshipment port",
    "LOADING PORT": "Port of Loading",
    "PORT OF LOADING": "Port of Loading",
    "ORIGIN PORT": "Port of Loading",
    "POD": "Port Of Discharge",
    "PORT OF DISCHARGE": "Port Of Discharge",
    "DESTINATION PORT": "Port Of Discharge",
    "PICK UP DATE AT ORIGIN": "Pick Up Date at Origin",
    "SHIPMENT DATE BASED ON PICK UP AT ORIGIN": "Pick Up Date at Origin",
    "DISPATCH DATE": "Pick Up Date at Origin",
    "1ST ETD": "First ETD",
    "FIRST ETD": "First ETD",
    "ORIGIN DEPARTURE PLANNED DATE ETD": "First ETD",
    "ORIGIN DEPARTURE ORIGINAL PLANNED DATE ETD": "Estimated Time of Departure",
    "EMPTY OUTGATED": "Empty Outgated",
    "FULL INGATED": "Full Ingated",
    "LOADED": "Loaded",
    "ESTIMATED TIME OF DEPARTURE": "Estimated Time of Departure",
    "ATD": "Vessel Departed",
    "VESSEL DEPARTED": "Vessel Departed",
    "ORIGIN DEPARTURE ACTUAL DATE": "Vessel Departed",
    "BOOKING ETA": "Booked ETA Port",
    "BOOKED ETA PORT": "Booked ETA Port",
    "DESTINATION ARRIVAL ORIGINAL PLANNED DATE ETA": "Booked ETA Port",
    "CURRENT ETA": "ETA To Port of Discharge",
    "ETA TO PORT OF DISCHARGE": "ETA To Port of Discharge",
    "DESTINATION ARRIVAL PLANNED DATE ETA": "ETA To Port of Discharge",
    "ATA DESTINATION": "Vessel Arrived",
    "VESSEL ARRIVED": "Vessel Arrived",
    "DESTINATION ARRIVAL ACTUAL DATE": "Vessel Arrived",
    "ESTIMATED TRANSIT TIME": "TRANSIT TIME",
    "TRANSIT TIME": "TRANSIT TIME",
    "VESSEL DISCHARGED": "Vessel Discharged",
    "CUSTOMS RELEASED": "Customs Released",
    "ETA TO FINAL RAIL FACILITY": "ETA To Final Rail Facility",
    "ETA TO FINAL TERMINAL": "ETA To Final Rail Facility",
    "FULL OUTGATED FROM RAIL PORT FACILITY": "Full Outgated From Rail/Port Facility",
    "LFD PORT RAMP": "LFD Port/Ramp",
    "LFD PORT/RAMP": "LFD Port/Ramp",
    "ETA DOOR": "ETA Door",
    "CONTAINER DELIVERED": "Container Delivered",
    "CARRIER": "Steamship Line",
    "CARRIER NAME": "Steamship Line",
    "STEAMSHIP LINE": "Steamship Line",
    "INVOICE NO": "Invoice #",
    "INVOICE #": "Invoice #",
    "REMARKS": "REMARKS",
    "STATUS": "Current Status",
    "CURRENT STATUS": "Current Status",
    "DLV DATE": "Dlv Date",
}

GOCOMET_TO_TEMPLATE = {
    "Reference": "PO#",
    "Container Number (If Ocean Shipment)": "Container Number",
    "Tracking Number": "Container Number",
    "Plant [Other Data]": "Delivery Location",
    "Vendor [Other Data]": "Shipper Name",
    "Carrier Name": "Steamship Line",
    "Origin Port": "Port of Loading",
    "Destination Port": "Port Of Discharge",
    "Dispatch Date": "Pick Up Date at Origin",
    "Status": "Current Status",
    "Created At": "Booking Request Date",
    "Origin Departure Vessel Name": "Vessel Name",
    "Origin Departure Original Planned Date (ETD)": "Estimated Time of Departure",
    "Origin Departure Planned Date (ETD)": "First ETD",
    "Origin Departure Actual Date": "Vessel Departed",
    "Destination Arrival Original Planned Date (ETA)": "Booked ETA Port",
    "Destination Arrival Planned Date (ETA)": "ETA To Port of Discharge",
    "Destination Arrival Actual Date": "Vessel Arrived",
    "Gate Out Actual Date": "Full Outgated From Rail/Port Facility",
}

MONTHS = {m.upper(): i for i, m in enumerate(["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"], start=1)}
MONTHS.update({"JANUARY":1,"FEBRUARY":2,"MARCH":3,"APRIL":4,"JUNE":6,"JULY":7,"AUGUST":8,"SEPTEMBER":9,"OCTOBER":10,"NOVEMBER":11,"DECEMBER":12,"SEPT":9})


def clean_header(x: Any) -> str:
    if x is None:
        return ""
    s = str(x).replace("\n", " ").replace("\r", " ").strip()
    s = re.sub(r"\s+", " ", s)
    return s


def norm_key(x: Any) -> str:
    s = clean_header(x).upper()
    s = re.sub(r"\[[^]]*\]", "", s)  # Plant [Other Data] -> Plant
    s = re.sub(r"\([^)]*\)", "", s)  # dates with parenthetical hints
    s = s.replace("#", " #")
    s = re.sub(r"[^A-Z0-9#]+", " ", s)
    return re.sub(r"\s+", " ", s).strip()


def is_blank(v: Any) -> bool:
    return v is None or str(v).strip().upper() in NA_VALUES


def text_clean(v: Any) -> str:
    if v is None:
        return NA
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    s = str(v).replace("\n", " ").replace("\r", " ").replace("\xa0", " ").strip()
    s = re.sub(r"\s+", " ", s)
    return s if s and s.upper() not in NA_VALUES else NA


def detect_source_type(path: Path) -> str:
    n = path.name.lower().replace("_", " ").replace("-", " ")
    n = " ".join(n.split())
    if "liyana" in n:
        return "liyana"
    if "scanglobal" in n or "scan global" in n:
        return "scanglobal"
    if "cargomar" in n:
        return "cargomar"
    if "sinpex" in n or "china shipment" in n:
        return "sinpex"
    if "gocomet" in n or "detailed" in n or "tracking" in n:
        return "gocomet"
    if "open order" in n:
        return "open_order"
    if "hotlist" in n or "hot list" in n or "corporate hot" in n:
        return "hotlist"
    return "unknown"


def find_input_files(input_dir: Path) -> Dict[str, List[Path]]:
    groups: Dict[str, List[Path]] = defaultdict(list)
    for p in input_dir.glob("*.xlsx"):
        if p.name.startswith("~$"):
            continue
        st = detect_source_type(p)
        groups[st].append(p)
    for k in groups:
        groups[k].sort(key=lambda x: x.stat().st_mtime, reverse=True)
    return groups


def excel_serial_to_date(v: float) -> Optional[date]:
    try:
        if v < 1 or v > 80000:
            return None
        return (datetime(1899, 12, 30) + timedelta(days=float(v))).date()
    except Exception:
        return None


def fix_year(y: int) -> int:
    # Handle typos like 12026 or 32026 by taking the final four digits.
    if y > 9999:
        tail = int(str(y)[-4:])
        if 1900 <= tail <= 2099:
            return tail
    if y < 100:
        return 2000 + y if y <= 49 else 1900 + y
    return y


def make_date(y: int, m: int, d: int) -> Optional[date]:
    try:
        return date(fix_year(int(y)), int(m), int(d))
    except Exception:
        return None


def maybe_correct_datetime_from_dayfirst_source(d: date, source_type: str) -> date:
    """
    Keep real Excel date objects exactly as Excel stores them.

    Important fix after v7/v8 testing:
    - The previous version tried to swap month/day for ambiguous Excel date objects.
    - That corrupted valid Excel dates in ScanGlobal/Liyana and caused AH/AI mismatches.
    - We should only apply day-first logic to TEXT dates from CargoMar/Sinpex/ScanGlobal.

    Example:
    - CargoMar text `7.2.2026` is parsed as DD.MM.YYYY -> 02/07/2026.
    - A real Excel date object `2026-07-02` remains 07/02/2026.
    """
    return d


def parse_numeric_date(a: str, b: str, y: str, sep: str, source_type: str) -> Optional[date]:
    n1, n2, yy = int(a), int(b), fix_year(int(y))

    # Source-specific rules for TEXT dates only.
    # CargoMar normally uses DD.MM.YYYY. Sinpex is mixed and its row-level
    # convention is detected before values reach this function.
    # ScanGlobal text dates are treated as DD/MM/YYYY.
    source_day_first = source_type in {
        "cargomar", "sinpex_day_first", "scanglobal"
    }

    # If one part is >12, there is only one valid interpretation.
    if n1 > 12 and n2 <= 12:
        return make_date(yy, n2, n1)  # DD/MM/YYYY
    if n2 > 12 and n1 <= 12:
        return make_date(yy, n1, n2)  # MM/DD/YYYY

    # Ambiguous values like 7.2.2026: use source rule.
    if source_day_first:
        return make_date(yy, n2, n1)

    # Unknown/GoComet/Open Order default to US month-first when ambiguous.
    return make_date(yy, n1, n2)


def infer_row_date_source_type(
    row: Tuple[Any, ...], colmap: List[Tuple[Optional[str], Optional[str]]], source_type: str
) -> str:
    """Infer mixed text-date order from unambiguous dates in one source row.

    Sinpex contains both MM.DD.YYYY and DD.MM.YYYY rows. CargoMar is normally
    day-first but also has a few month-first imported rows. A value such as
    07.02.2026 is ambiguous on its own, so use other date fields in the same
    shipment row (for example 07.15.2026 or 20.05.2026) to determine its order.
    """
    if source_type not in {"sinpex", "cargomar"}:
        return source_type

    day_first_evidence = 0
    month_first_evidence = 0
    for index, (_, target) in enumerate(colmap):
        if not target or not is_date_header(target) or index >= len(row):
            continue
        value = row[index]
        if is_blank(value) or isinstance(value, (date, datetime, int, float)):
            continue
        for match in re.finditer(
            r"\b(\d{1,2})([./-])(\d{1,2})\2(\d{2,5})\b", text_clean(value)
        ):
            first, second = int(match.group(1)), int(match.group(3))
            if first > 12 and second <= 12:
                day_first_evidence += 1
            elif second > 12 and first <= 12:
                month_first_evidence += 1

    if day_first_evidence > month_first_evidence:
        return f"{source_type}_day_first"
    if month_first_evidence > day_first_evidence:
        return f"{source_type}_month_first"
    # Current Sinpex rows are predominantly US month-first; CargoMar's native
    # convention remains day-first when a row contains only ambiguous dates.
    return "sinpex_month_first" if source_type == "sinpex" else "cargomar"


def parse_date_value(value: Any, source_type: str = "unknown") -> Any:
    """Return date object where possible. If a cell contains multiple dates, return the latest date."""
    if is_blank(value):
        return NA
    if isinstance(value, datetime):
        return maybe_correct_datetime_from_dayfirst_source(value.date(), source_type)
    if isinstance(value, date):
        return maybe_correct_datetime_from_dayfirst_source(value, source_type)
    if isinstance(value, (int, float)):
        d = excel_serial_to_date(float(value))
        return d if d else value

    s = text_clean(value)
    if s == NA:
        return NA
    s = s.strip().strip(".")
    found: List[date] = []

    # 28-Apr-2026 or 28 April 2026
    for m in re.finditer(r"\b(\d{1,2})[-\s]([A-Za-z]{3,9})[-\s](\d{2,4})\b", s):
        mon = MONTHS.get(m.group(2).upper())
        if mon:
            d = make_date(int(m.group(3)), mon, int(m.group(1)))
            if d:
                found.append(d)

    # ISO 2026-07-05 / 2026/07/05
    for m in re.finditer(r"\b(20\d{2})[-/\.](\d{1,2})[-/\.](\d{1,2})\b", s):
        d = make_date(int(m.group(1)), int(m.group(2)), int(m.group(3)))
        if d:
            found.append(d)

    # Numeric dates: 15.05.2026, 15/05/2026, 16-04-2026, 05-28-26
    for m in re.finditer(r"\b(\d{1,2})([./-])(\d{1,2})\2(\d{2,5})\b", s):
        d = parse_numeric_date(m.group(1), m.group(3), m.group(4), m.group(2), source_type)
        if d:
            found.append(d)

    if found:
        return max(found)
    # Preserve malformed date-like text so Date_Format_Check can report it.
    if re.search(r"\d{1,2}[./-]\d{1,2}[./-]\d{2,5}", s):
        return s
    return s


def is_date_header(header: str) -> bool:
    h = norm_key(header)
    return any(k in h for k in DATE_KEYWORDS)


def extract_po_tokens(value: Any) -> List[str]:
    if is_blank(value):
        return []
    s = text_clean(value).upper()
    # PO numbers are usually 5+ digits. This avoids capturing invoice suffixes like 23/24.
    tokens = re.findall(r"\b\d{5,}\b", s)
    out = []
    seen = set()
    for t in tokens:
        if t not in seen:
            out.append(t); seen.add(t)
    return out


def extract_container_tokens(value: Any) -> List[str]:
    if is_blank(value):
        return []
    s = text_clean(value).upper().replace(" ", "")
    # Standard container format 4 letters + 7 digits.
    tokens = re.findall(r"\b[A-Z]{4}\d{7}\b", s)
    out = []
    seen = set()
    for t in tokens:
        if t not in seen:
            out.append(t); seen.add(t)
    if out:
        return out
    # fallback split for odd container-looking values in forwarder files only
    parts = re.split(r"[,/;|]+", text_clean(value).upper())
    for p in parts:
        p = re.sub(r"[^A-Z0-9]", "", p.strip())
        if p and p != NA and len(p) >= 8 and p not in seen:
            out.append(p); seen.add(p)
    return out


def extract_standard_container_tokens(value: Any) -> List[str]:
    """Strict container extraction: only ISO style 4 letters + 7 digits. Used for GoComet and Hotlist.

    Uses the original text with word boundaries so it does not accidentally turn FEDEX numbers
    into fake containers (example: FEDEX 495516593016 must not become EDEX4955165).
    """
    if is_blank(value):
        return []
    s = text_clean(value).upper()
    raw_tokens = re.findall(r"\b[A-Z]{4}\s*\d{7}\b", s)
    out = []
    seen = set()
    for t in raw_tokens:
        t = re.sub(r"[^A-Z0-9]", "", t)
        if t not in seen:
            out.append(t); seen.add(t)
    return out


def join_unique(values: Iterable[Any]) -> str:
    out = []
    seen = set()
    for v in values:
        if is_blank(v):
            continue
        s = text_clean(v)
        if s == NA:
            continue
        if s not in seen:
            out.append(s); seen.add(s)
    return ", ".join(out) if out else NA


def merge_po_strings(existing: Any, new_refs: Iterable[Any]) -> str:
    tokens = extract_po_tokens(existing)
    for r in new_refs:
        tokens.extend(extract_po_tokens(r))
    out = []
    seen = set()
    for t in tokens:
        if t not in seen:
            out.append(t); seen.add(t)
    return ",".join(out) if out else (text_clean(existing) if not is_blank(existing) else NA)


def normalize_delivery(value: Any) -> str:
    s = text_clean(value).upper()
    if s in {NA, "30"}:
        return NA
    compact = re.sub(r"[^A-Z0-9]+", " ", s).strip()
    # Specific rules first
    if any(x in compact for x in ["GEORGIA", "RINCON", "GTC"]):
        return "GTC"
    if any(x in compact for x in ["POCATELLO", "POCATOELLO", "VTCU"]):
        return "VTCU"
    if any(x in compact for x in ["CHIHUAHUA", "MEXICO", "VTC WEST", "VTCW", "VTV MEXICO", "WEST S A DE C V", "WEST SA DE CV"]):
        return "VTCW"
    if any(x in compact for x in ["LIVINGSTON", "EL PASO"]):
        return "EL PASO, TX"
    if any(x in compact for x in [
        "BLUE HILLS", "TROUTVILLE", "ROANOKE", "ROANOKE VA",
        "RONAOKE", "ROANOE", "ROANAOKE", "SIEBEL",
        "VIRGINIA TRANSFORMER", "VTCR",
    ]):
        return "VTCR"
    if re.fullmatch(r"VTC\.?", compact):
        return "VTCR"
    return s


def normalize_incoterm(value: Any) -> str:
    s = text_clean(value).upper()
    if s == NA:
        return NA
    for code in ["EXW", "DAP", "CFR", "FCA", "FOB"]:
        if re.search(rf"\b{code}\b", s):
            return code
    return s


def normalize_shipper(value: Any) -> str:
    s = text_clean(value).upper()
    if s == NA:
        return NA
    s = re.sub(r"\s+", " ", s).strip(" ,")
    # Remove obvious duplicated legal punctuation, keep company name readable.
    replacements = {
        "CO., LTD.": "CO LTD", "CO.,LTD": "CO LTD", "CO. LTD.": "CO LTD", "LTD.": "LTD",
        "PRIVATE LIMITED": "PVT LTD", "PVT. LTD.": "PVT LTD", "PVT LTD.": "PVT LTD",
    }
    for a,b in replacements.items():
        s = s.replace(a, b)
    return re.sub(r"\s+", " ", s).strip(" ,")


def normalize_steamship(value: Any) -> str:
    s = text_clean(value).upper()
    if s == NA:
        return NA
    if "HAPAG" in s:
        return "HAPAG-LLOYD"
    if "EVERGREEN" in s or s == "EMC":
        return "EVERGREEN"
    if "COSCO" in s:
        return "COSCO"
    if "OOCL" in s:
        return "OOCL"
    if "ONE" in s or "OCEAN NETWORK" in s:
        return "ONE"
    if "MAERSK" in s:
        return "MAERSK"
    if "MSC" in s:
        return "MSC"
    if "CMA" in s:
        return "CMA CGM"
    return s


def normalize_port(value: Any) -> str:
    s = text_clean(value).upper()
    if s == NA:
        return NA
    s = s.replace(";", ",").replace("/", " / ")
    s = re.sub(r"\([^)]*\)", "", s)  # remove port/terminal code in parentheses
    s = s.replace("UNITED STATES", "USA").replace("AMERICA", "USA")
    s = re.sub(r"\s*,\s*", ", ", s)
    s = re.sub(r"\s+", " ", s).strip(" ,")
    # remove trailing UNLOC-like extra code after country/city
    s = re.sub(r",\s*[A-Z]{2,}[A-Z0-9]{3,}\s*$", "", s)
    key = re.sub(r"[^A-Z0-9]+", " ", s).strip()

    rules = [
        (["NHAVA SHEVA", "NHAVA SHVA", "JAWAHARLAL NEHRU", "JNPT", "NHAVA"], "NHAVA SHEVA, INDIA"),
        (["MUNDRA"], "MUNDRA, INDIA"),
        (["CHENNAI"], "CHENNAI, INDIA"),
        (["BANGALORE ICD", "BANGALORE"], "BANGALORE ICD, INDIA"),
        (["BARODA"], "BARODA, INDIA"),
        (["ENNORE"], "ENNORE, INDIA"),
        (["MUMBAI", "SAHAR"], "MUMBAI, INDIA"),
        (["SHANGHAI"], "SHANGHAI, CHINA"),
        (["NINGBO"], "NINGBO, CHINA"),
        (["QINGDAO"], "QINGDAO, CHINA"),
        (["DALIAN"], "DALIAN, CHINA"),
        (["CHANGZHOU"], "CHANGZHOU, CHINA"),
        (["GUANGZHOU"], "GUANGZHOU, CHINA"),
        (["HOUSTON"], "HOUSTON, TX"),
        (["NORFOLK"], "NORFOLK, VA"),
        (["SAVANNAH"], "SAVANNAH, GA"),
        (["LONG BEACH", "LGB"], "LONG BEACH, CA"),
        (["LOS ANGELES", "LAX", " LA "], "LOS ANGELES, CA"),
        (["NEW YORK"], "NEW YORK, NY"),
        (["NEWARK"], "NEWARK, NJ"),
        (["FREEPORT"], "FREEPORT, USA"),
        (["TACOMA"], "TACOMA, WA"),
        (["SEATTLE"], "SEATTLE, WA"),
        (["MANZANILLO"], "MANZANILLO, MEXICO"),
        (["GENOA"], "GENOA, ITALY"),
        (["LIVORNO"], "LIVORNO, ITALY"),
        (["GEBZE"], "GEBZE, TURKEY"),
        (["AMBARLI"], "AMBARLI, TURKEY"),
        (["IZMIT"], "IZMIT, TURKEY"),
        (["IZMIR"], "IZMIR, TURKEY"),
        (["ADANA"], "ADANA, TURKEY"),
        (["EVYAP"], "EVYAPAN, TURKEY"),
        (["HAMBURG"], "HAMBURG, GERMANY"),
        (["BREMERHAVEN"], "BREMERHAVEN, GERMANY"),
        (["JEBEL ALI"], "JEBEL ALI, UNITED ARAB EMIRATES"),
        (["DUBAI"], "DUBAI, UNITED ARAB EMIRATES"),
        (["SOHAR"], "SOHAR, OMAN"),
        (["SALALAH"], "SALALAH, OMAN"),
        (["COLOMBO"], "COLOMBO, SRI LANKA"),
        (["VALENCIA"], "VALENCIA, SPAIN"),
        (["YANTIAN"], "YANTIAN, CHINA"),
        (["BUSAN"], "BUSAN, SOUTH KOREA"),
        (["GREENSBORO"], "GREENSBORO, NC"),
        (["ATLANTA"], "ATLANTA, GA"),
        (["SALT LAKE CITY"], "SALT LAKE CITY, UT"),
        (["CHARLESTON"], "CHARLESTON, SC"),
    ]
    padded = f" {key} "
    for terms, out in rules:
        for term in terms:
            term_key = re.sub(r"[^A-Z0-9]+", " ", term).strip()
            if term_key == "LA":
                if re.search(r"\bLA\b", padded):
                    return out
            elif term_key and term_key in key:
                return out

    # Standardize common city/state forms without adding too many assumptions.
    city_state = {
        r"^HOUSTON TX$": "HOUSTON, TX", r"^NORFOLK VA$": "NORFOLK, VA", r"^SAVANNAH GA$": "SAVANNAH, GA",
        r"^LONG BEACH CA$": "LONG BEACH, CA", r"^LOS ANGELES CA$": "LOS ANGELES, CA", r"^NEW YORK NY$": "NEW YORK, NY",
    }
    for pat, out in city_state.items():
        if re.match(pat, key):
            return out
    return s


def make_unique_headers(headers: List[Any]) -> List[Optional[str]]:
    counts = defaultdict(int)
    out: List[Optional[str]] = []
    for h in headers:
        h2 = clean_header(h)
        if not h2:
            out.append(None)
            continue
        counts[h2] += 1
        out.append(h2 if counts[h2] == 1 else f"{h2}_{counts[h2]}")
    return out


def iter_nonblank_rows(ws, start_row: int = 2, max_blank_rows: int = MAX_TRAILING_BLANK_ROWS):
    blanks = 0
    for idx, row in enumerate(ws.iter_rows(min_row=start_row, values_only=True), start=start_row):
        if not any(not is_blank(v) for v in row):
            blanks += 1
            if blanks >= max_blank_rows:
                break
            continue
        blanks = 0
        yield idx, row


def relevant_sheet_names(wb, source_type: str) -> List[str]:
    if source_type == "open_order":
        return [OPEN_ORDER_SHEET_NAME] if OPEN_ORDER_SHEET_NAME in wb.sheetnames else [wb.sheetnames[0]]
    if source_type == "gocomet":
        for nm in wb.sheetnames:
            if "tracking" in nm.lower():
                return [nm]
        return [wb.sheetnames[0]]
    if source_type == "liyana":
        return wb.sheetnames  # VTC DSR, China Sea, Dubai Sea
    return [wb.sheetnames[0]]


def get_base_headers(liyana_path: Path) -> List[str]:
    wb = open_input_workbook(liyana_path)
    sheet = "VTC DSR 2026" if "VTC DSR 2026" in wb.sheetnames else wb.sheetnames[0]
    ws = wb[sheet]
    headers = [clean_header(c) for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True)) if clean_header(c)]
    return ["Lane number" if h == "Liyana Lane number" else h for h in headers]


def map_header(source_header: str, occurrence: int) -> Optional[str]:
    if not source_header:
        return None
    nk = norm_key(source_header)
    # Repeated Voy columns from CargoMar/Sinpex
    if nk == "VOY" and occurrence == 2:
        return "Voy_2"
    if nk == "VOY" and occurrence >= 3:
        return "Voy_3"
    # Repeated vessel columns already have names Second Vessel/Final Vessel and are preserved in extras.
    return ALIASES.get(nk, source_header if source_header in EXTRA_HEADERS_KEEP else None)


def normalize_value(header: str, value: Any, source_type: str) -> Any:
    if is_blank(value):
        return NA
    if is_date_header(header):
        return parse_date_value(value, source_type)
    if header == "Delivery Location":
        return normalize_delivery(value)
    if header in {"Transshipment port", "Port of Loading", "Port Of Discharge"}:
        return normalize_port(value)
    if header == "Incoterm":
        return normalize_incoterm(value)
    if header == "Shipper Name":
        return normalize_shipper(value)
    if header == "Steamship Line":
        return normalize_steamship(value)
    if header == "PO#":
        toks = extract_po_tokens(value)
        return ",".join(toks) if toks else text_clean(value)
    return text_clean(value)


def read_forwarder_rows(paths: List[Path], base_headers: List[str], final_headers: List[str], stats: Dict[str, Any]) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    for path in paths:
        source_type = detect_source_type(path)
        wb = open_input_workbook(path)
        for sheet_name in relevant_sheet_names(wb, source_type):
            ws = wb[sheet_name]
            raw_headers = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
            unique_headers = make_unique_headers(raw_headers)
            occurrence = defaultdict(int)
            colmap: List[Tuple[Optional[str], Optional[str]]] = []
            for raw, unique in zip(raw_headers, unique_headers):
                if unique is None:
                    colmap.append((None, None)); continue
                occurrence[norm_key(raw)] += 1
                target = map_header(unique if unique in EXTRA_HEADERS_KEEP else clean_header(raw), occurrence[norm_key(raw)])
                colmap.append((unique, target))

            included = 0
            for excel_row, row in iter_nonblank_rows(ws, 2):
                row_source_type = infer_row_date_source_type(row, colmap, source_type)
                out = {h: NA for h in final_headers}
                out["Source File"] = path.name
                out["Source Sheet"] = sheet_name
                out["Source Row"] = excel_row
                for i, (_, target) in enumerate(colmap):
                    if not target or target not in out:
                        continue
                    val = row[i] if i < len(row) else None
                    if is_blank(val):
                        continue
                    cleaned = normalize_value(target, val, row_source_type)
                    if out.get(target, NA) == NA:
                        out[target] = cleaned
                rows.append(out)
                included += 1
            stats["source_rows"].append([path.name, sheet_name, source_type, included])
    return rows


def build_gocomet_index(paths: List[Path], stats: Dict[str, Any]) -> Tuple[Dict[str, Dict[str, Any]], List[Dict[str, Any]]]:
    """
    Build a container-level index from GoComet / Detailed Tracking.

    Business rule:
    - Do NOT add raw GoComet-only columns.
    - Use only fields that map to the Liyana template.
    - Match by container number.
    - Bring Reference into PO#.
    - Add GoComet rows only when their container is missing from the forwarder consolidated data.

    Important v10 fix:
    - We do not filter by Mode anymore. Some GoComet files use blank/non-standard Mode values.
    - Instead, a GoComet row is considered in scope if it has a valid container number
      in Tracking Number or Container Number (If Ocean Shipment).
    """
    index: Dict[str, Dict[str, Any]] = {}
    raw_records: List[Dict[str, Any]] = []
    stats["gocomet_source_rows_scanned"] = 0
    stats["gocomet_valid_container_rows"] = 0
    stats["gocomet_rows_without_valid_container"] = 0
    stats["gocomet_rows_added_to_index"] = 0

    for path in paths:
        wb = open_input_workbook(path)
        for sheet_name in relevant_sheet_names(wb, "gocomet"):
            ws = wb[sheet_name]
            headers = [clean_header(h) for h in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
            hidx = {h: i for i, h in enumerate(headers) if h}
            count = 0
            valid_count = 0
            no_container_count = 0

            for excel_row, row in iter_nonblank_rows(ws, 2):
                count += 1
                stats["gocomet_source_rows_scanned"] += 1

                keys: Set[str] = set()
                for col in ["Container Number (If Ocean Shipment)", "Tracking Number", "Container Number", "Container No"]:
                    if col in hidx and hidx[col] < len(row):
                        keys.update(extract_standard_container_tokens(row[hidx[col]]))

                if not keys:
                    no_container_count += 1
                    stats["gocomet_rows_without_valid_container"] += 1
                    continue

                valid_count += 1
                stats["gocomet_valid_container_rows"] += 1

                rec: Dict[str, Any] = {"Source File": "Detailed Tracking", "Source Sheet": sheet_name, "Source Row": excel_row}
                for src, target in GOCOMET_TO_TEMPLATE.items():
                    if src not in hidx:
                        continue
                    val = row[hidx[src]] if hidx[src] < len(row) else None
                    if is_blank(val):
                        continue
                    rec[target] = normalize_value(target, val, "gocomet")

                ref = row[hidx["Reference"]] if "Reference" in hidx and hidx["Reference"] < len(row) else None
                rec["PO#"] = merge_po_strings(rec.get("PO#", NA), [ref])

                for k in keys:
                    existing = index.get(k, {})
                    merged = dict(existing)
                    for a, b in rec.items():
                        if a == "PO#":
                            merged[a] = merge_po_strings(merged.get(a, NA), [b])
                        elif is_blank(merged.get(a, NA)) and not is_blank(b):
                            merged[a] = b
                    merged["Container Number"] = k
                    index[k] = merged
                    stats["gocomet_rows_added_to_index"] += 1

                raw_records.append({**rec, "_keys": sorted(keys)})

            stats["source_rows"].append([path.name, sheet_name, "gocomet_scanned", count])
            stats["source_rows"].append([path.name, sheet_name, "gocomet_valid_container_rows", valid_count])
            if no_container_count:
                stats["warnings"].append(f"GoComet file {path.name} sheet {sheet_name}: {no_container_count} rows skipped because no valid container number was found.")

    stats["gocomet_rows"] += stats.get("gocomet_source_rows_scanned", 0)
    stats["gocomet_containers"] = len(index)
    return index, raw_records

def apply_gocomet(rows: List[Dict[str, Any]], gidx: Dict[str, Dict[str, Any]], final_headers: List[str], stats: Dict[str, Any]) -> List[Dict[str, Any]]:
    existing_keys: Set[str] = set()
    updated_po = 0
    fields_filled = 0
    for row in rows:
        keys = extract_container_tokens(row.get("Container Number"))
        for k in keys:
            existing_keys.add(k)
            grecord = gidx.get(k)
            if not grecord:
                continue
            before = row.get("PO#", NA)
            row["PO#"] = merge_po_strings(row.get("PO#", NA), [grecord.get("PO#", NA)])
            if row.get("PO#", NA) != before:
                updated_po += 1
            for h in final_headers:
                if h in {"Source File", "Source Sheet", "Source Row", "PO#"}:
                    continue
                gv = grecord.get(h, NA)
                if h in grecord and not is_blank(gv) and is_blank(row.get(h, NA)):
                    row[h] = gv
                    fields_filled += 1
    # Add GoComet containers not found in forwarder files, but only mapped template fields.
    added = 0
    for k, grecord in gidx.items():
        if k in existing_keys:
            continue
        out = {h: NA for h in final_headers}
        out["Source File"] = "Detailed Tracking"
        out["Source Sheet"] = grecord.get("Source Sheet", "Detailed Tracking")
        out["Source Row"] = grecord.get("Source Row", NA)
        for h in final_headers:
            if h in grecord and h not in {"Source File", "Source Sheet", "Source Row"}:
                out[h] = grecord[h]
        out["Container Number"] = k
        rows.append(out)
        added += 1
    stats["gocomet_po_rows_updated"] = updated_po
    stats["gocomet_fields_filled"] = fields_filled
    stats["gocomet_missing_container_rows_added"] = added
    return rows


def build_open_order_index(paths: List[Path], stats: Dict[str, Any]) -> Dict[str, Dict[str, Any]]:
    index: Dict[str, Dict[str, Any]] = defaultdict(lambda: {"Dlv Date": [], "Buyer": [], "Vendor": [], "Class": []})
    for path in paths:
        wb = open_input_workbook(path)
        sheet = OPEN_ORDER_SHEET_NAME if OPEN_ORDER_SHEET_NAME in wb.sheetnames else wb.sheetnames[0]
        ws = wb[sheet]
        headers = [clean_header(h) for h in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
        hidx = {h: i for i, h in enumerate(headers) if h}
        po_col = None
        for candidate in ["PO Nbr", "PO#", "PO Number", "PO No", "PO"]:
            if candidate in hidx:
                po_col = hidx[candidate]
                break
        if po_col is None:
            stats["warnings"].append(f"Open Order file {path.name}: PO column not found")
            continue
        count = 0
        for _, row in iter_nonblank_rows(ws, 2, max_blank_rows=1000):
            po_raw = row[po_col] if po_col < len(row) else None
            pos = extract_po_tokens(po_raw)
            if not pos:
                continue
            dlv = row[hidx["Dlv Date"]] if "Dlv Date" in hidx and hidx["Dlv Date"] < len(row) else None
            buyer = row[hidx["Buyer"]] if "Buyer" in hidx and hidx["Buyer"] < len(row) else None
            vendor = None
            if "Vendor Name" in hidx and hidx["Vendor Name"] < len(row):
                vendor = row[hidx["Vendor Name"]]
            elif "Vendor" in hidx and hidx["Vendor"] < len(row):
                vendor = row[hidx["Vendor"]]
            cls = row[hidx["Class"]] if "Class" in hidx and hidx["Class"] < len(row) else None
            for po in pos:
                index[po]["Dlv Date"].append(parse_date_value(dlv, "open_order"))
                index[po]["Buyer"].append(text_clean(buyer))
                index[po]["Vendor"].append(text_clean(vendor))
                index[po]["Class"].append(text_clean(cls))
            count += 1
        stats["open_order_rows"] += count
    # Collapse
    collapsed: Dict[str, Dict[str, Any]] = {}
    for po, d in index.items():
        dates = [x for x in d["Dlv Date"] if isinstance(x, date)]
        collapsed[po] = {
            "Dlv Date": min(dates) if dates else NA,  # earliest due date is safest for delivery requirement
            "Buyer": join_unique(d["Buyer"]),
            "Vendor": join_unique(d["Vendor"]),
            "Class": join_unique(d["Class"]),
        }
    stats["open_order_po_count"] = len(collapsed)
    return collapsed


def apply_open_order(rows: List[Dict[str, Any]], oo: Dict[str, Dict[str, Any]], stats: Dict[str, Any]) -> None:
    matched_rows = 0
    matched_po = 0
    for row in rows:
        pos = extract_po_tokens(row.get("PO#"))
        recs = [oo[p] for p in pos if p in oo]
        if not recs:
            continue
        matched_rows += 1
        matched_po += len(recs)
        # Dlv Date: earliest among all matched POs
        dates = [r["Dlv Date"] for r in recs if isinstance(r.get("Dlv Date"), date)]
        if dates:
            row["Dlv Date"] = min(dates)
        for field in ["Buyer", "Vendor", "Class"]:
            row[field] = join_unique(r.get(field) for r in recs)
    stats["open_order_rows_matched"] = matched_rows
    stats["open_order_po_matches"] = matched_po


def build_container_po_mapping(rows: List[Dict[str, Any]], oo: Dict[str, Dict[str, Any]]) -> List[List[Any]]:
    table = [["Container Number", "PO#", "Dlv Date", "Buyer", "Vendor", "Class"]]
    seen = set()
    for row in rows:
        containers = extract_container_tokens(row.get("Container Number"))
        pos = extract_po_tokens(row.get("PO#"))
        for c in containers:
            for p in pos:
                key = (c, p)
                if key in seen:
                    continue
                seen.add(key)
                rec = oo.get(p, {})
                table.append([c, p, rec.get("Dlv Date", NA), rec.get("Buyer", NA), rec.get("Vendor", NA), rec.get("Class", NA)])
    return table


def final_date_cleanup(rows: List[Dict[str, Any]], final_headers: List[str], stats: Dict[str, Any]) -> None:
    text_dates_left = 0
    date_issues: List[List[Any]] = []
    for row_num, row in enumerate(rows, start=2):
        for h in final_headers:
            if is_date_header(h):
                before = row.get(h, NA)
                after = parse_date_value(before, "unknown")
                row[h] = after
                if isinstance(after, str) and re.search(r"\d{1,2}[./-]\d{1,2}[./-]\d{2,5}", after):
                    text_dates_left += 1
                    if len(date_issues) < 5000:
                        date_issues.append([
                            row_num, h, text_clean(before), text_clean(after),
                            row.get("Source File", NA), row.get("Source Sheet", NA), row.get("Source Row", NA),
                            "Date-like text remained after parsing"
                        ])
    stats["text_date_values_left"] = text_dates_left
    stats["date_format_issues"] = date_issues


def dedupe_rows(rows: List[Dict[str, Any]], stats: Dict[str, Any]) -> List[Dict[str, Any]]:
    # Do not drop rows aggressively because one shipment can have multiple POs/containers. Only remove exact duplicate rows.
    seen = set()
    out = []
    for r in rows:
        key = tuple(text_clean(r.get(h)) for h in ["Source File", "Source Sheet", "Source Row", "Container Number", "PO#", "Steamship Line Bill Of Lading #"])
        if key in seen:
            continue
        seen.add(key)
        out.append(r)
    stats["exact_duplicate_rows_removed"] = len(rows) - len(out)
    return out



CONTAINER_REGEX = re.compile(r"\b[A-Z]{4}\s*\d{7}\b", re.IGNORECASE)


def normalize_container_key(value: Any) -> str:
    if value is None:
        return ""
    return re.sub(r"[^A-Z0-9]", "", str(value).upper().strip())


def extract_hotlist_containers_from_text(text: Any) -> Set[str]:
    """Extract normalized ISO-style container numbers from transportation text."""
    if is_blank(text):
        return set()
    return {normalize_container_key(m) for m in extract_standard_container_tokens(text)}


def is_transportation_header(value: Any) -> bool:
    """Recognize the Hot List transportation header despite spacing variations.

    Ignore long narrative cells that happen to contain the words transport/tracking
    (seen on the PD Tracker sheet).
    """
    text = text_clean(value)
    if not text or len(text) > 80:
        return False
    key = norm_key(text)
    return "TRANSPORT" in key and ("TRACK" in key or "INFO" in key)


def build_hotlist_index(paths: List[Path], stats: Dict[str, Any]) -> Set[str]:
    """
    Read all Corporate Hot List sheets and extract containers from Transportation / Tracking Info.

    The source workbook uses columns N–P, and the active column can change between sections
    in the same sheet. After any transportation header is seen, scan every candidate column
    on each row — do not lock to only the most recent header column (that missed containers
    still written in the previous column, e.g. ONEU2378629 on VTCW P2 Med Power).
    """
    hotlist_containers: Set[str] = set()
    match_details: List[List[Any]] = []
    rows_scanned = 0
    max_hotlist_row = 10000
    first_candidate_col = 14  # N
    last_candidate_col = 16   # P

    for path in paths:
        wb = open_input_workbook(path)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            section_started = False
            max_row = min(ws.max_row or 1, max_hotlist_row)
            for row_idx, cells in enumerate(
                ws.iter_rows(
                    min_row=1,
                    max_row=max_row,
                    min_col=first_candidate_col,
                    max_col=last_candidate_col,
                    values_only=True,
                ),
                start=1,
            ):
                for cell_value in cells:
                    if is_transportation_header(cell_value):
                        section_started = True

                if not section_started:
                    continue

                row_had_transport_text = False
                for value in cells:
                    if is_blank(value) or is_transportation_header(value):
                        continue
                    found = extract_hotlist_containers_from_text(value)
                    if not found:
                        continue
                    row_had_transport_text = True
                    for container in sorted(found):
                        hotlist_containers.add(container)
                        match_details.append([
                            path.name,
                            sheet_name,
                            row_idx,
                            container,
                            text_clean(value),
                        ])
                if row_had_transport_text:
                    rows_scanned += 1

        close_input_workbook(wb)

    stats["hotlist_files"] = len(paths)
    stats["hotlist_rows_scanned"] = rows_scanned
    stats["hotlist_unique_containers"] = len(hotlist_containers)
    stats["hotlist_match_details"] = match_details
    return hotlist_containers

def apply_hotlist(rows: List[Dict[str, Any]], hotlist_containers: Set[str], stats: Dict[str, Any]) -> None:
    """Mark Hotlist = Yes when any consolidated container matches the Hot List container set; otherwise No."""
    yes = 0
    no = 0
    matched: Set[str] = set()
    for row in rows:
        containers = {normalize_container_key(c) for c in extract_container_tokens(row.get("Container Number"))}
        hit = containers.intersection(hotlist_containers) if containers else set()
        if hit:
            row["Hotlist"] = "Yes"
            yes += 1
            matched.update(hit)
        else:
            row["Hotlist"] = "No"
            no += 1
    unmatched = sorted(hotlist_containers - matched)
    stats["hotlist_rows_yes"] = yes
    stats["hotlist_rows_no"] = no
    stats["hotlist_containers_matched_in_ocean"] = len(matched)
    stats["hotlist_containers_not_in_ocean"] = unmatched
    if unmatched:
        stats.setdefault("warnings", []).append(
            "Hotlist containers not found in ocean DSR files (cannot mark on dashboard): "
            + ", ".join(unmatched)
        )


def audit_hotlist_open_order(rows: List[Dict[str, Any]], oo: Dict[str, Dict[str, Any]], stats: Dict[str, Any]) -> None:
    """
    Validate that Hotlist=Yes rows received Dlv Date/Buyer/Vendor/Class from Open Order when possible.
    If a hotlist row has no PO# or a PO is not present in Open Order, we report it in an audit sheet
    rather than silently pretending the row is complete.
    """
    audit_rows: List[List[Any]] = []
    yes_rows = 0
    complete_rows = 0
    missing_dlv = 0
    missing_po = 0
    po_not_in_open_order = 0
    for row in rows:
        if text_clean(row.get("Hotlist")) != "Yes":
            continue
        yes_rows += 1
        containers = ",".join(extract_container_tokens(row.get("Container Number"))) or NA
        po_text = row.get("PO#", NA)
        pos = extract_po_tokens(po_text)
        dlv = row.get("Dlv Date", NA)
        issue = "OK"
        if not pos:
            issue = "Missing PO# - cannot lookup Open Order"
            missing_po += 1
        else:
            missing = [p for p in pos if p not in oo]
            if missing:
                issue = "PO not found in Open Order: " + ",".join(missing)
                po_not_in_open_order += 1
        if is_blank(dlv):
            if issue == "OK":
                issue = "Dlv Date missing after Open Order match"
            missing_dlv += 1
        else:
            if issue == "OK":
                complete_rows += 1
        audit_rows.append([
            containers,
            text_clean(po_text),
            dlv,
            row.get("Buyer", NA),
            row.get("Vendor", NA),
            row.get("Class", NA),
            issue,
            row.get("Source File", NA),
            row.get("Source Sheet", NA),
            row.get("Source Row", NA),
        ])
    stats["hotlist_open_order_audit"] = audit_rows
    stats["hotlist_yes_rows_checked_for_open_order"] = yes_rows
    stats["hotlist_yes_complete_with_dlv_date"] = complete_rows
    stats["hotlist_yes_missing_dlv_date"] = missing_dlv
    stats["hotlist_yes_missing_po"] = missing_po
    stats["hotlist_yes_po_not_in_open_order"] = po_not_in_open_order


def write_cell(ws, row_idx: int, col_idx: int, value: Any, date_format, text_format) -> None:
    """xlsxwriter uses zero-based row/col indexes."""
    if is_blank(value):
        ws.write(row_idx, col_idx, NA, text_format)
    elif isinstance(value, datetime):
        ws.write_datetime(row_idx, col_idx, value, date_format)
    elif isinstance(value, date):
        ws.write_datetime(row_idx, col_idx, datetime(value.year, value.month, value.day), date_format)
    else:
        ws.write(row_idx, col_idx, value, text_format)


def _write_workbook(rows: List[Dict[str, Any]], final_headers: List[str], container_po_table: List[List[Any]], stats: Dict[str, Any], out_path: Path) -> None:
    """Fast xlsxwriter output with real Excel dates and mm/dd/yyyy format."""
    out_path.parent.mkdir(parents=True, exist_ok=True)
    with xlsxwriter.Workbook(str(out_path), {"constant_memory": True, "strings_to_urls": False, "strings_to_formulas": False}) as wb:
        header_fmt = wb.add_format({"bold": True, "font_color": "white", "bg_color": "#1F4E78", "align": "center", "valign": "vcenter", "text_wrap": True})
        date_fmt = wb.add_format({"num_format": DATE_FORMAT})
        text_fmt = wb.add_format({})
    
        ws = wb.add_worksheet(OUTPUT_SHEET_NAME)
        for c, h in enumerate(final_headers):
            ws.write(0, c, h, header_fmt)
        date_flags = [is_date_header(h) for h in final_headers]
        for r_idx, row in enumerate(rows, start=1):
            for c_idx, h in enumerate(final_headers):
                write_cell(ws, r_idx, c_idx, row.get(h, NA), date_fmt if date_flags[c_idx] else text_fmt, text_fmt)
        ws.freeze_panes(1, 0)
        ws.autofilter(0, 0, len(rows), len(final_headers) - 1)
        for i, h in enumerate(final_headers):
            width = 16
            if i <= 2:
                width = 22
            if h == "Source File":
                width = 42
            if h in {"Shipper Name", "Consignee Name", "Delivery Location", "PO#", "REMARKS", "Current Status", "Vendor"}:
                width = 28
            ws.set_column(i, i, width)
    
        if CREATE_AUDIT_SHEETS:
            mapws = wb.add_worksheet("Container_PO_Mapping")
            for c, h in enumerate(container_po_table[0]):
                mapws.write(0, c, h, header_fmt)
            map_date_flags = [is_date_header(h) for h in container_po_table[0]]
            for r_idx, r in enumerate(container_po_table[1:], start=1):
                for c_idx, v in enumerate(r):
                    write_cell(mapws, r_idx, c_idx, v, date_fmt if map_date_flags[c_idx] else text_fmt, text_fmt)
            mapws.freeze_panes(1, 0)
            mapws.autofilter(0, 0, max(1, len(container_po_table) - 1), len(container_po_table[0]) - 1)
            for c in range(len(container_po_table[0])):
                mapws.set_column(c, c, 22)
    
            sumws = wb.add_worksheet("Summary")
            summary_rows = [
                ["Metric", "Value"],
                ["Output rows", len(rows)],
                ["Output columns", len(final_headers)],
                ["Container PO mapping rows", max(0, len(container_po_table) - 1)],
                ["GoComet source rows scanned", stats.get("gocomet_source_rows_scanned", 0)],
                ["GoComet valid container rows", stats.get("gocomet_valid_container_rows", 0)],
                ["GoComet rows without valid container", stats.get("gocomet_rows_without_valid_container", 0)],
                ["GoComet containers indexed", stats.get("gocomet_containers", 0)],
                ["GoComet missing container rows added", stats.get("gocomet_missing_container_rows_added", 0)],
                ["Rows with PO updated from GoComet", stats.get("gocomet_po_rows_updated", 0)],
                ["Template fields filled from GoComet", stats.get("gocomet_fields_filled", 0)],
                ["Open Order PO count", stats.get("open_order_po_count", 0)],
                ["Rows matched to Open Order", stats.get("open_order_rows_matched", 0)],
                ["Text date values left in date columns", stats.get("text_date_values_left", 0)],
                ["Exact duplicate rows removed", stats.get("exact_duplicate_rows_removed", 0)],
                ["Hotlist files found", stats.get("hotlist_files", 0)],
                ["Hotlist unique containers found", stats.get("hotlist_unique_containers", 0)],
                ["Hotlist containers matched in ocean", stats.get("hotlist_containers_matched_in_ocean", 0)],
                ["Hotlist containers not in ocean", len(stats.get("hotlist_containers_not_in_ocean", []) or [])],
                ["Rows marked Hotlist = Yes", stats.get("hotlist_rows_yes", 0)],
                ["Rows marked Hotlist = No", stats.get("hotlist_rows_no", 0)],
                ["Hotlist Yes rows checked for Open Order", stats.get("hotlist_yes_rows_checked_for_open_order", 0)],
                ["Hotlist Yes with Dlv Date", stats.get("hotlist_yes_complete_with_dlv_date", 0)],
                ["Hotlist Yes missing Dlv Date", stats.get("hotlist_yes_missing_dlv_date", 0)],
                ["Hotlist Yes missing PO#", stats.get("hotlist_yes_missing_po", 0)],
                ["Hotlist Yes PO not found in Open Order", stats.get("hotlist_yes_po_not_in_open_order", 0)],
            ]
            for r_idx, r in enumerate(summary_rows):
                for c_idx, v in enumerate(r):
                    sumws.write(r_idx, c_idx, v, header_fmt if r_idx == 0 else text_fmt)
            sumws.set_column(0, 0, 42)
            sumws.set_column(1, 1, 25)
    
    
            hotmatches = stats.get("hotlist_match_details", [])
            hotws = wb.add_worksheet("Hotlist_Matches")
            hot_headers = ["Hotlist File", "Hotlist Sheet", "Hotlist Row", "Container Number", "Transportation / Tracking Info"]
            for c, h in enumerate(hot_headers):
                hotws.write(0, c, h, header_fmt)
            for r_idx, rec in enumerate(hotmatches, start=1):
                for c_idx, v in enumerate(rec):
                    hotws.write(r_idx, c_idx, v, text_fmt)
            hotws.freeze_panes(1, 0)
            hotws.autofilter(0, 0, max(1, len(hotmatches)), len(hot_headers) - 1)
            hotws.set_column(0, 0, 42)
            hotws.set_column(1, 1, 22)
            hotws.set_column(2, 3, 18)
            hotws.set_column(4, 4, 70)
    
            hoo_rows = stats.get("hotlist_open_order_audit", [])
            hoows = wb.add_worksheet("Hotlist_OpenOrder_Check")
            hoo_headers = ["Container Number", "PO#", "Dlv Date", "Buyer", "Vendor", "Class", "Issue", "Source File", "Source Sheet", "Source Row"]
            for c, h in enumerate(hoo_headers):
                hoows.write(0, c, h, header_fmt)
            hoo_date_flags = [is_date_header(h) for h in hoo_headers]
            for r_idx, rec in enumerate(hoo_rows, start=1):
                for c_idx, v in enumerate(rec):
                    write_cell(hoows, r_idx, c_idx, v, date_fmt if hoo_date_flags[c_idx] else text_fmt, text_fmt)
            hoows.freeze_panes(1, 0)
            hoows.autofilter(0, 0, max(1, len(hoo_rows)), len(hoo_headers) - 1)
            hoows.set_column(0, 1, 24)
            hoows.set_column(2, 2, 14)
            hoows.set_column(3, 5, 24)
            hoows.set_column(6, 6, 42)
            hoows.set_column(7, 9, 22)
    
            date_issues = stats.get("date_format_issues", [])
            datews = wb.add_worksheet("Date_Format_Check")
            date_headers = ["Output Row", "Column", "Original Value", "Parsed/Remaining Value", "Source File", "Source Sheet", "Source Row", "Issue"]
            for c, h in enumerate(date_headers):
                datews.write(0, c, h, header_fmt)
            for r_idx, rec in enumerate(date_issues, start=1):
                for c_idx, v in enumerate(rec):
                    datews.write(r_idx, c_idx, v, text_fmt)
            if not date_issues:
                datews.write(1, 0, "No date-like text values remained in date columns.", text_fmt)
            datews.freeze_panes(1, 0)
            datews.autofilter(0, 0, max(1, len(date_issues)), len(date_headers) - 1)
            datews.set_column(0, 0, 14)
            datews.set_column(1, 1, 34)
            datews.set_column(2, 3, 26)
            datews.set_column(4, 4, 42)
            datews.set_column(5, 7, 24)
    
            inclusion_ws = wb.add_worksheet("Data_Inclusion_Audit")
            inc_headers = ["File", "Sheet", "Source/Step", "Rows Counted"]
            for c, h in enumerate(inc_headers):
                inclusion_ws.write(0, c, h, header_fmt)
            for r_idx, rec in enumerate(stats.get("source_rows", []), start=1):
                for c_idx, v in enumerate(rec):
                    inclusion_ws.write(r_idx, c_idx, v, text_fmt)
            inclusion_ws.freeze_panes(1, 0)
            inclusion_ws.autofilter(0, 0, max(1, len(stats.get("source_rows", []))), len(inc_headers) - 1)
            inclusion_ws.set_column(0, 0, 42)
            inclusion_ws.set_column(1, 2, 24)
            inclusion_ws.set_column(3, 3, 18)
    
            valws = wb.add_worksheet("Validation_Report")
            val_headers = ["Type", "File", "Sheet", "Source", "Rows/Message"]
            for c, h in enumerate(val_headers):
                valws.write(0, c, h, header_fmt)
            r_idx = 1
            for rec in stats.get("source_rows", []):
                vals = ["Source rows", rec[0], rec[1], rec[2], rec[3]]
                for c_idx, v in enumerate(vals):
                    valws.write(r_idx, c_idx, v, text_fmt)
                r_idx += 1
            for w in stats.get("warnings", []):
                vals = ["Warning", "", "", "", w]
                for c_idx, v in enumerate(vals):
                    valws.write(r_idx, c_idx, v, text_fmt)
                r_idx += 1
            if stats.get("text_date_values_left", 0):
                vals = ["Warning", "", "", "", f"{stats['text_date_values_left']} date-like text values remained. Check source formats."]
                for c_idx, v in enumerate(vals):
                    valws.write(r_idx, c_idx, v, text_fmt)
            valws.freeze_panes(1, 0)
            valws.autofilter(0, 0, max(1, r_idx), len(val_headers) - 1)
            valws.set_column(0, 0, 20)
            valws.set_column(1, 1, 42)
            valws.set_column(2, 3, 22)
            valws.set_column(4, 4, 60)
    


def validate_output_workbook(path: Path, expected_rows: int, expected_headers: List[str]) -> None:
    """Reject an incomplete or structurally invalid workbook before publication."""
    with zipfile.ZipFile(path, "r") as archive:
        bad_member = archive.testzip()
        if bad_member:
            raise ValueError(f"Corrupt XLSX member: {bad_member}")
        ET.fromstring(archive.read("xl/workbook.xml"))

    wb = open_input_workbook(path)
    try:
        if OUTPUT_SHEET_NAME not in wb.sheetnames:
            raise ValueError(f"Missing output sheet: {OUTPUT_SHEET_NAME}")
        ws = wb[OUTPUT_SHEET_NAME]
        actual_headers = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
        if actual_headers != expected_headers:
            raise ValueError("Output header validation failed")
        if ws.max_row != expected_rows + 1:
            raise ValueError(
                f"Output row validation failed: expected {expected_rows}, found {ws.max_row - 1}"
            )
    finally:
        close_input_workbook(wb)


def write_workbook(rows: List[Dict[str, Any]], final_headers: List[str], container_po_table: List[List[Any]], stats: Dict[str, Any], out_path: Path) -> None:
    """Create, validate, and atomically publish the production workbook."""
    out_path.parent.mkdir(parents=True, exist_ok=True)
    fd, temporary_name = tempfile.mkstemp(
        prefix=f".{out_path.stem}.", suffix=".tmp.xlsx", dir=out_path.parent
    )
    os.close(fd)
    temporary_path = Path(temporary_name)
    try:
        _write_workbook(rows, final_headers, container_po_table, stats, temporary_path)
        validate_output_workbook(temporary_path, len(rows), final_headers)
        os.replace(temporary_path, out_path)
    finally:
        temporary_path.unlink(missing_ok=True)


def _main() -> int:
    input_dir = INPUT_DIR
    output_dir = OUTPUT_DIR
    out_path = output_dir / OUTPUT_FILE
    stats: Dict[str, Any] = defaultdict(int)
    stats["source_rows"] = []
    stats["warnings"] = []

    if not input_dir.exists():
        print(f"ERROR: Input folder does not exist: {input_dir}")
        return 2
    groups = find_input_files(input_dir)
    # Both files are essential to the consolidated dataset. Refuse to replace
    # the production output when either is missing; otherwise an incomplete
    # workbook can look like a successful refresh.
    needed = ["liyana", "gocomet"]
    for n in needed:
        if not groups.get(n):
            print(f"ERROR: Missing required file type: {n}. Put the file in {input_dir}")
            return 2

    liyana_path = groups["liyana"][0]
    base_headers = get_base_headers(liyana_path)
    final_headers = ["Source File", "Source Sheet", "Source Row"] + base_headers
    for h in EXTRA_HEADERS_KEEP + OPEN_ORDER_OUTPUT_COLUMNS + ["Hotlist"]:
        if h not in final_headers:
            final_headers.append(h)

    forwarder_paths = []
    for key in ["scanglobal", "liyana", "cargomar", "sinpex"]:
        forwarder_paths.extend(groups.get(key, []))
    rows = read_forwarder_rows(forwarder_paths, base_headers, final_headers, stats)

    gidx, _ = build_gocomet_index(groups.get("gocomet", []), stats)
    if gidx:
        rows = apply_gocomet(rows, gidx, final_headers, stats)

    # Build Hotlist BEFORE Open Order because the Open Order workbook is large and can increase memory.
    # This keeps the Hot List step fast/reliable even when the Corporate Hot List has heavily formatted tabs.
    hotlist_containers = build_hotlist_index(groups.get("hotlist", []), stats)

    oo = build_open_order_index(groups.get("open_order", []), stats)
    if oo:
        apply_open_order(rows, oo, stats)

    if hotlist_containers:
        apply_hotlist(rows, hotlist_containers, stats)
    else:
        for row in rows:
            row["Hotlist"] = "No"
        if not groups.get("hotlist"):
            stats["warnings"].append("No Corporate Hot List file found in input folder. Hotlist column set to No for all rows.")

    audit_hotlist_open_order(rows, oo, stats)
    final_date_cleanup(rows, final_headers, stats)
    rows = dedupe_rows(rows, stats)
    container_po = build_container_po_mapping(rows, oo)

    write_workbook(rows, final_headers, container_po, stats, out_path)
    print(f"DONE: {out_path}")
    print(f"Rows: {len(rows):,}")
    print(f"Columns: {len(final_headers):,}")
    print(f"Container_PO_Mapping rows: {max(0, len(container_po)-1):,}")
    print(f"Text-date values left: {stats.get('text_date_values_left', 0)}")
    return 0


def main() -> int:
    """Run the ETL and guarantee cleanup of all input workbooks."""
    try:
        return _main()
    finally:
        close_all_input_workbooks()


if __name__ == "__main__":
    raise SystemExit(main())





